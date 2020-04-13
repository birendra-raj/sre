import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
wbook=Workbook()
wbook.save("D:\\Selenium_with_python\\PLearn\\Lesson\\demo.xlsx")

from openpyxl import load_workbook
filepath="D:\\Selenium_with_python\\PLearn\\Lesson\\demo.xlsx"
wbook=load_workbook(filepath)
wsheet=wbook.active

wsheet['A1']='Birendra'
wsheet.cell(row=1,column=2).value='Churendra'
wbook.save(filepath)

